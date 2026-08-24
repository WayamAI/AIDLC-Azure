"""Write AIDLC live smoke-test results to Excel."""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[2] / "AIDLC_Live_Test_Report.xlsx"

# Results from live_smoke run on 2026-08-23 (kimi-k3:cloud intended; AI path still hit llama3.1 until restart)
ROWS = [
    # case_id, feature, method, endpoint, expected, actual_status, result, detail, notes
    (1, "Platform", "GET", "/health", 200, 200, "PASS", '{"status":"healthy"}', "Backend process healthy"),
    (2, "Auth", "GET", "/api/auth/status", 200, 200, "PASS", "workos=false, password_auth=true, dev_login=true", "Password auth enabled"),
    (3, "Auth", "POST", "/api/auth/login", 200, 200, "PASS", "Logged in as mriganka.dey@wayam.ai", "Seed account wayam/wayam"),
    (4, "Auth", "GET", "/api/auth/me", 200, 200, "PASS", "email=mriganka.dey@wayam.ai org=Wayam", "Session cookie works"),
    (5, "Dashboard", "GET", "/api/dashboard/stats", 200, 200, "PASS", "stats returned (empty counters)", "Dashboard API OK"),
    (6, "Requirements", "GET", "/api/requirements", 200, 200, "PASS", "count=0", "List OK, empty DB"),
    (7, "Requirements + AI", "POST", "/api/requirements", "200/201", 500, "FAIL", "model 'llama3.1' not found", "Backend still used llama3.1 at request time; .env now kimi-k3:cloud needs process restart"),
    (8, "Test Cases", "GET", "/api/test-cases", 200, 200, "PASS", "len=0", "List OK"),
    (9, "Synthetic Data", "GET", "/api/synthetic-data", 200, 200, "PASS", "[]", "List OK"),
    (10, "Prioritization", "GET", "/api/prioritization", 200, 200, "PASS", "[]", "List OK"),
    (11, "Pipeline", "GET", "/api/pipeline/runs", 200, 200, "PASS", "[]", "List OK"),
    (12, "Deployments", "GET", "/api/deployments/health", 200, 200, "PASS", "configured=false", "Vercel not configured expected"),
    (13, "Root Cause", "GET", "/api/testing/root-cause", 200, 200, "PASS", "summary + items=[]", "Feature wired"),
    (14, "Root Cause", "GET", "/api/testing/root-cause/failures", 200, 200, "PASS", "failures=[]", "Feature wired"),
    (15, "Test Selection", "GET", "/api/testing/test-selection/history", 200, 200, "PASS", "runs=[]", "Feature wired"),
    (16, "Self-Healing", "GET", "/api/testing/healing", 200, 200, "PASS", "summary + items=[]", "Feature wired"),
    (17, "Live Testing", "GET", "/api/repo/runs", 200, 200, "PASS", "runs=[]", "Feature wired"),
    (18, "Incidents", "GET", "/api/incidents", 200, 200, "PASS", "[]", "List OK"),
    (19, "Cost Tracker", "GET", "/api/cost-logs", 200, 200, "PASS", "logs=[] total=0", "List OK"),
    (20, "GitHub", "GET", "/api/github/repo-info", 200, 422, "FAIL", "missing query owner/repo", "Endpoint requires owner+repo query params smoke call incomplete, not a product outage"),
    (21, "Jira", "GET", "/api/jira/projects", "200 or config error", 400, "PASS*", "Jira URL missing http(s) protocol", "Integration not configured fails gracefully"),
    (22, "Monitoring", "POST", "/api/monitoring/simulate-time-series", 200, 500, "FAIL", "Internal Server Error", "Simulate endpoint crashed"),
    (23, "Release Gate", "POST", "/api/release-gate/evaluate", "200/400/422", "conn reset", "FAIL", "WinError 10054 connection closed", "Server connection dropped mid-request"),
    (24, "Organizations", "GET", "/api/orgs/current", 200, 200, "PASS", "Wayam org returned", "Multi-tenant org OK"),
    (25, "Frontend", "GET", "http://127.0.0.1:8081/", 200, 200, "PASS", "HTML shell served", "Vite dev server OK"),
    (26, "AI Infra", "POST", "http://localhost:11434/api/chat", 200, 200, "PASS", "kimi-k3:cloud replied READY", "Kimi K3 cloud reachable via local Ollama"),
    (27, "Auth", "POST", "/api/auth/signup", 409, 409, "PASS", "duplicate email rejected", "Idempotent seed protection"),
    (28, "Auth", "POST", "/api/auth/logout", 200, 200, "PASS", "logged_out", "Logout clears session"),
    (29, "Auth", "GET", "/api/auth/me (after logout)", 401, 401, "PASS", "unauthorized", "Session invalidated"),
    (30, "Auth", "POST", "/api/auth/login (bad password)", 401, 401, "PASS", "Invalid email or password", "Auth hardening OK"),
]

SUMMARY_NOTES = [
    ("Model configured in .env", "kimi-k3:cloud"),
    ("Model used by AI generate at test time", "llama3.1 (stale process) → FAIL case 7"),
    ("Direct Ollama kimi-k3:cloud", "WORKING"),
    ("Seed login", "mriganka.dey@wayam.ai / wayam"),
    ("SESSION_SECRET", "wayam"),
    ("Frontend", "http://localhost:8081"),
    ("Backend", "http://127.0.0.1:8000"),
    ("Pass", "26 (incl. 1 PASS* for expected Jira misconfig)"),
    ("Fail", "4 AI model mismatch, GitHub smoke params, Monitoring 500, Release gate conn reset"),
    ("Generated at (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
]


def style_header(cell, fill_hex: str) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=fill_hex)
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def main() -> None:
    wb = Workbook()

    # ── Sheet 1: Results
    ws = wb.active
    ws.title = "Live Test Results"
    headers = [
        "Case #",
        "Feature Area",
        "HTTP Method",
        "Endpoint / Target",
        "Expected",
        "Actual Status",
        "Result",
        "Response / Detail",
        "Notes",
    ]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        style_header(ws.cell(1, col), "1F2937")

    fill_pass = PatternFill("solid", fgColor="DCFCE7")
    fill_fail = PatternFill("solid", fgColor="FEE2E2")
    fill_soft = PatternFill("solid", fgColor="FEF9C3")
    thin = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    for row in ROWS:
        ws.append(list(row))
        r_idx = ws.max_row
        result = row[6]
        fill = fill_pass if result.startswith("PASS") else fill_fail
        if result == "PASS*":
            fill = fill_soft
        for c in range(1, 10):
            cell = ws.cell(r_idx, c)
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c == 7:
                cell.fill = fill
                cell.font = Font(bold=True)

    widths = [8, 18, 12, 42, 18, 14, 10, 55, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:I{ws.max_row}"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    # ── Sheet 2: Summary
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Item", "Value"])
    style_header(ws2.cell(1, 1), "1F2937")
    style_header(ws2.cell(1, 2), "1F2937")
    for item, value in SUMMARY_NOTES:
        ws2.append([item, value])
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 70

    # ── Sheet 3: Working vs Not
    ws3 = wb.create_sheet("Working vs Not")
    ws3.append(["Status", "Feature / Check", "Evidence"])
    for col in range(1, 4):
        style_header(ws3.cell(1, col), "1F2937")

    working = [
        ("WORKING", "Health / Auth / Session", "Login, me, logout, bad password, duplicate signup"),
        ("WORKING", "Dashboard stats", "200 OK"),
        ("WORKING", "Requirements list", "200 OK (empty)"),
        ("WORKING", "Test cases / Synthetic / Prioritization", "200 OK lists"),
        ("WORKING", "Pipeline runs", "200 OK"),
        ("WORKING", "Deployments health endpoint", "200 OK (Vercel not configured)"),
        ("WORKING", "Root Cause APIs", "list + failures"),
        ("WORKING", "Test Selection history", "200 OK"),
        ("WORKING", "Self-Healing list", "200 OK"),
        ("WORKING", "Repo runs / Incidents / Cost logs", "200 OK"),
        ("WORKING", "Organizations current", "Wayam org"),
        ("WORKING", "Frontend Vite :8081", "200 OK"),
        ("WORKING", "Kimi K3 via Ollama cloud", "Direct chat READY"),
        ("NOT WORKING", "AI requirement generation through AIDLC API", "Called llama3.1 (missing) set kimi-k3:cloud, restart backend"),
        ("NOT WORKING", "Monitoring simulate-time-series", "HTTP 500"),
        ("NOT WORKING", "Release gate evaluate (this call)", "Connection reset"),
        ("PARTIAL", "GitHub repo-info", "Needs owner+repo query; not a hard outage"),
        ("PARTIAL", "Jira projects", "Credentials/URL not configured"),
        ("PARTIAL", "Vercel deployments", "Token/project not configured"),
    ]
    for status, feature, evidence in working:
        ws3.append([status, feature, evidence])
        r = ws3.max_row
        if status == "WORKING":
            ws3.cell(r, 1).fill = fill_pass
        elif status == "NOT WORKING":
            ws3.cell(r, 1).fill = fill_fail
        else:
            ws3.cell(r, 1).fill = fill_soft
    for i, w in enumerate([14, 42, 70], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
